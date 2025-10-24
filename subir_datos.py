# subir_datos.py
# Lógica para subir y aplicar mapeos de productos desde CSV/XLSX.
# Columnas requeridas (case-insensitive): SKU, nombre, bodega

from __future__ import annotations
import io
import csv
from typing import Any, Dict, List, Tuple, Iterable
import persistencia_postgresql as db

# Intentar soportar .xlsx sin obligar a pandas.
try:
    import openpyxl  # para leer .xlsx
except Exception:
    openpyxl = None

def _norm(s: Any) -> str:
    return (str(s).strip() if s is not None else "")

def _leer_csv(buf: bytes) -> List[Dict[str, Any]]:
    filas: List[Dict[str, Any]] = []
    txt = buf.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(txt))
    for raw in reader:
        filas.append({k: raw.get(k) for k in raw.keys()})
    return filas

def _leer_xlsx(buf: bytes) -> List[Dict[str, Any]]:
    if openpyxl is None:
        raise RuntimeError("No se puede leer .xlsx: instala openpyxl o sube un CSV.")
    wb = openpyxl.load_workbook(io.BytesIO(buf), data_only=True)
    ws = wb.active  # primera hoja
    # Encabezados
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(_norm(cell.value))
    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2):
        d = {}
        for i, cell in enumerate(row):
            key = headers[i] if i < len(headers) else f"col{i+1}"
            d[key] = cell.value
        rows.append(d)
    return rows

def _estandarizar_columnas(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Acepta headers en cualquier combinación de mayúsculas/minúsculas/espacios:
    SKU / sku / Sku, nombre / Nombre, bodega / BODEGA, alias / Alias
    """
    salida: List[Dict[str, Any]] = []
    def keymap(k: str) -> str:
        k2 = (k or "").strip().lower()
        if k2 in ("sku", "código", "codigo", "itemcode", "item"):
            return "sku"
        if k2 in ("nombre", "name", "descripcion", "descripción"):
            return "nombre"
        if k2 in ("bodega", "whscode", "warehouse"):
            return "bodega"
        if k2 in ("alias", "aliases", "apodo", "apodos"):
            return "alias"
        return k2  # se ignora lo demás
    for r in rows:
        nr = { keymap(k): (r.get(k)) for k in r.keys() }
        # normalizar strings
        nr["sku"]    = _norm(nr.get("sku"))
        nr["nombre"] = _norm(nr.get("nombre"))
        nr["bodega"] = _norm(nr.get("bodega"))
        nr["alias"]  = _norm(nr.get("alias"))
        salida.append(nr)
    return salida

def _buscar_o_crear_producto(cur, sku: str, nombre: str) -> int | None:
    """
    Busca un producto existente por SKU. Si no existe, lo crea.
    """
    if not (sku or nombre):
        return None
    
    # Buscar producto existente por SKU
    if sku:
        cur.execute("SELECT id FROM productos WHERE sku = %s", (sku,))
        row = cur.fetchone()
        if row:
            return int(row[0])
    
    # Si no existe, crear nuevo producto
    cur.execute("""
        INSERT INTO productos (sku, nombre, activo)
        VALUES (%s, %s, TRUE)
        RETURNING id;
    """, (sku or "", nombre or ""))
    return int(cur.fetchone()[0])

def _agregar_alias_producto(cur, producto_id: int, cliente_id: int, alias: str) -> bool:
    """
    Agrega un alias a un producto si no existe ya.
    Retorna True si se agregó, False si ya existía.
    """
    if not alias or not alias.strip():
        return False
    
    alias_clean = alias.strip()
    
    # Verificar si el alias ya existe
    cur.execute("""
        SELECT id FROM producto_alias 
        WHERE producto_id = %s AND cliente_id = %s AND alias = %s
    """, (producto_id, cliente_id, alias_clean))
    
    if cur.fetchone():
        print(f"DEBUG: Alias '{alias_clean}' ya existe para producto {producto_id}")
        return False
    
    # Insertar el alias
    try:
        cur.execute("""
            INSERT INTO producto_alias (producto_id, cliente_id, alias)
            VALUES (%s, %s, %s)
        """, (producto_id, cliente_id, alias_clean))
        print(f"DEBUG: Alias '{alias_clean}' agregado exitosamente para producto {producto_id}")
        return True
    except Exception as e:
        print(f"DEBUG: Error al insertar alias '{alias_clean}': {str(e)}")
        return False

def _upsert_bodega_por_cliente(conn, cliente_id: int, filas: Iterable[Dict[str,str]]) -> Tuple[int,int,int,List[str]]:
    ins = act = om = 0
    errores: List[str] = []
    alias_agregados = 0
    
    with conn.cursor() as cur:
        for i, r in enumerate(filas, start=2):  # +2 por encabezado
            try:
                sku, nombre, bodega, alias = r.get("sku",""), r.get("nombre",""), r.get("bodega",""), r.get("alias","")
                
                if not (sku or nombre):
                    om += 1; continue
                if not bodega:
                    errores.append(f"Fila {i}: bodega vacía.")
                    om += 1; continue
                
                # Buscar o crear producto (evita duplicados por SKU)
                pid = _buscar_o_crear_producto(cur, sku, nombre)
                if not pid:
                    errores.append(f"Fila {i}: no se pudo crear/obtener producto (SKU='{sku}', nombre='{nombre}').")
                    om += 1; continue
                
                # Agregar alias si existe
                if alias and alias.strip():
                    print(f"DEBUG: Agregando alias '{alias}' para producto ID {pid}, cliente ID {cliente_id}")
                    if _agregar_alias_producto(cur, pid, cliente_id, alias):
                        alias_agregados += 1
                        print(f"DEBUG: Alias agregado exitosamente")
                    else:
                        print(f"DEBUG: Alias ya existía o hubo error")
                
                # update si existe, sino insert
                cur.execute("""
                    SELECT id FROM bodegas_producto_por_cliente
                    WHERE cliente_id = %s AND producto_id = %s;
                """, (cliente_id, pid))
                row = cur.fetchone()
                if row:
                    cur.execute("""
                        UPDATE bodegas_producto_por_cliente
                           SET bodega = %s
                         WHERE id = %s;
                    """, (bodega, int(row[0])))
                    act += 1
                else:
                    cur.execute("""
                        INSERT INTO bodegas_producto_por_cliente (cliente_id, producto_id, bodega)
                        VALUES (%s, %s, %s);
                    """, (cliente_id, pid, bodega))
                    ins += 1
                    
            except Exception as e:
                errores.append(f"Fila {i}: Error procesando fila - {str(e)}")
                om += 1
                continue
    
    # Agregar información de alias al resultado
    if alias_agregados > 0:
        errores.append(f"INFO: Se agregaron {alias_agregados} alias nuevos.")
    
    return ins, act, om, errores

def _upsert_bodega_por_sucursal(conn, sucursal_id: int, filas: Iterable[Dict[str,str]]) -> Tuple[int,int,int,List[str]]:
    ins = act = om = 0
    errores: List[str] = []
    alias_agregados = 0
    
    with conn.cursor() as cur:
        # Verificar que la sucursal existe
        cur.execute("SELECT cliente_id FROM sucursales WHERE id = %s;", (sucursal_id,))
        cliente_result = cur.fetchone()
        if not cliente_result:
            raise RuntimeError("Sucursal no encontrada.")
        
        cliente_id = cliente_result[0]
        
        for i, r in enumerate(filas, start=2):
            try:
                sku, nombre, bodega, alias = r.get("sku",""), r.get("nombre",""), r.get("bodega",""), r.get("alias","")
                
                if not (sku or nombre):
                    om += 1; continue
                if not bodega:
                    errores.append(f"Fila {i}: bodega vacía.")
                    om += 1; continue
                
                # Buscar o crear producto (evita duplicados por SKU)
                pid = _buscar_o_crear_producto(cur, sku, nombre)
                if not pid:
                    errores.append(f"Fila {i}: no se pudo crear/obtener producto (SKU='{sku}', nombre='{nombre}').")
                    om += 1; continue
                
                # Agregar alias si existe
                if alias and alias.strip():
                    print(f"DEBUG: Agregando alias '{alias}' para producto ID {pid}, cliente ID {cliente_id}")
                    if _agregar_alias_producto(cur, pid, cliente_id, alias):
                        alias_agregados += 1
                        print(f"DEBUG: Alias agregado exitosamente")
                    else:
                        print(f"DEBUG: Alias ya existía o hubo error")
                
                # update/insert
                cur.execute("""
                    SELECT id FROM bodegas_producto_por_sucursal
                    WHERE sucursal_id = %s AND producto_id = %s;
                """, (sucursal_id, pid))
                row = cur.fetchone()
                if row:
                    cur.execute("""
                        UPDATE bodegas_producto_por_sucursal
                           SET bodega = %s
                         WHERE id = %s;
                    """, (bodega, int(row[0])))
                    act += 1
                else:
                    cur.execute("""
                        INSERT INTO bodegas_producto_por_sucursal (sucursal_id, producto_id, bodega)
                        VALUES (%s, %s, %s);
                    """, (sucursal_id, pid, bodega))
                    ins += 1
                    
            except Exception as e:
                errores.append(f"Fila {i}: Error procesando fila - {str(e)}")
                om += 1
                continue
    
    # Agregar información de alias al resultado
    if alias_agregados > 0:
        errores.append(f"INFO: Se agregaron {alias_agregados} alias nuevos.")
    
    return ins, act, om, errores

def _verificar_modalidad_cliente(cur, cliente_id: int) -> bool:
    cur.execute("SELECT COALESCE(usa_bodega_por_sucursal, FALSE) FROM clientes WHERE id = %s;", (cliente_id,))
    row = cur.fetchone()
    if not row:
        raise RuntimeError("cliente_id no existe.")
    return bool(row[0])

def _validar_sucursal_pertenece(cur, sucursal_id: int, cliente_id: int) -> None:
    cur.execute("SELECT 1 FROM sucursales WHERE id = %s AND cliente_id = %s AND activo = TRUE;", (sucursal_id, cliente_id))
    if not cur.fetchone():
        raise RuntimeError("La sucursal no pertenece al cliente o está inactiva.")

def cargar_y_aplicar_mapeos_productos(archivo_bytes: bytes, nombre_archivo: str,
                                      cliente_id: int, sucursal_id: int | None) -> Dict[str, Any]:
    if not archivo_bytes:
        raise RuntimeError("Archivo vacío.")

    nombre = (nombre_archivo or "").lower()
    if nombre.endswith(".csv"):
        rows = _leer_csv(archivo_bytes)
    elif nombre.endswith(".xlsx"):
        rows = _leer_xlsx(archivo_bytes)
    else:
        raise RuntimeError("Formato no soportado. Usa .csv o .xlsx")

    filas = _estandarizar_columnas(rows)

    with db.obtener_conexion() as conn:
        with conn.cursor() as cur:
            por_sucursal = _verificar_modalidad_cliente(cur, cliente_id)
            if por_sucursal:
                if not sucursal_id:
                    raise RuntimeError("Este cliente usa bodega por sucursal: debes seleccionar la sucursal.")
                _validar_sucursal_pertenece(cur, sucursal_id, cliente_id)
            else:
                if sucursal_id:
                    # ignoramos sucursal si la mandan por error
                    sucursal_id = None

        # Ejecutar upsert según modalidad
        if por_sucursal:
            ins, act, om, errs = _upsert_bodega_por_sucursal(conn, sucursal_id, filas)
        else:
            ins, act, om, errs = _upsert_bodega_por_cliente(conn, cliente_id, filas)

    return {
        "insertados": ins,
        "actualizados": act,
        "omitidos": om,
        "errores": errs,
        "modo": "por_sucursal" if sucursal_id else "por_cliente"
    }

def _estandarizar_columnas_sucursales(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Acepta headers en cualquier combinación de mayúsculas/minúsculas/espacios:
    SAP / sap, Alias / alias, Encargado / encargado, Direccion / direccion, Telefono / telefono, RUC / ruc, Ciudad / ciudad
    """
    salida: List[Dict[str, Any]] = []
    def keymap(k: str) -> str:
        k2 = (k or "").strip().lower()
        if k2 in ("sap", "almacen", "warehouse", "whscode"):
            return "sap"
        if k2 in ("alias", "nombre_alias", "alias_pdf"):
            return "alias"
        if k2 in ("encargado", "responsable", "contacto"):
            return "encargado"
        if k2 in ("direccion", "dirección", "address", "ubicacion", "ubicación"):
            return "direccion"
        if k2 in ("ruc", "tax_id", "nit"):
            return "ruc"
        if k2 in ("ciudad", "city", "localidad"):
            return "ciudad"
        return k2  # se ignora lo demás
    for r in rows:
        nr = { keymap(k): (r.get(k)) for k in r.keys() }
        # normalizar strings
        nr["sap"] = _norm(nr.get("sap"))
        nr["alias"] = _norm(nr.get("alias"))
        nr["encargado"] = _norm(nr.get("encargado"))
        nr["direccion"] = _norm(nr.get("direccion"))
        nr["ruc"] = _norm(nr.get("ruc"))
        nr["ciudad"] = _norm(nr.get("ciudad"))
        salida.append(nr)
    return salida

def _upsert_sucursal(conn, cliente_id: int, filas: Iterable[Dict[str,str]]) -> Tuple[int,int,int,List[str]]:
    ins = act = om = 0
    errores: List[str] = []
    with conn.cursor() as cur:
        for i, r in enumerate(filas, start=2):  # +2 por encabezado
            sap, alias, encargado, direccion, ruc, ciudad = (
                r.get("sap",""), r.get("alias",""), r.get("encargado",""), 
                r.get("direccion",""), r.get("ruc",""), r.get("ciudad","")
            )
            if not sap:
                errores.append(f"Fila {i}: SAP vacío.")
                om += 1; continue
            
            # Limitar SAP a 10 caracteres para el campo almacen
            sap_almacen = sap[:10] if len(sap) > 10 else sap
            
            # El nombre de la sucursal siempre es el valor completo del campo SAP
            nombre_sucursal = sap
            
            # Siempre insertar nueva sucursal (permitir todos los duplicados)
            cur.execute("""
                INSERT INTO sucursales (cliente_id, almacen, alias, nombre, encargado, direccion, ruc, ciudad)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s);
            """, (cliente_id, sap_almacen, alias, nombre_sucursal, encargado, direccion, ruc, ciudad))
            ins += 1
    
    return ins, act, om, errores

def cargar_y_aplicar_mapeos_sucursales(archivo_bytes: bytes, nombre_archivo: str,
                                      cliente_id: int) -> Dict[str, Any]:
    if not archivo_bytes:
        raise RuntimeError("Archivo vacío.")

    nombre = (nombre_archivo or "").lower()
    if nombre.endswith(".csv"):
        rows = _leer_csv(archivo_bytes)
    elif nombre.endswith(".xlsx"):
        rows = _leer_xlsx(archivo_bytes)
    else:
        raise RuntimeError("Formato no soportado. Usa .csv o .xlsx")

    filas = _estandarizar_columnas_sucursales(rows)

    with db.obtener_conexion() as conn:
        # Ejecutar upsert de sucursales
        ins, act, om, errs = _upsert_sucursal(conn, cliente_id, filas)

    return {
        "insertados": ins,
        "actualizados": act,
        "omitidos": om,
        "errores": errs,
        "modo": "sucursales"
    }
